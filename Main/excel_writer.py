import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def save_videos_to_excel(videos, filename='youtube_trending.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YouTube Trending"

    headers = ['ลำดับ', 'ชื่อวิดีโอ', 'ช่อง', 'ยอดวิว', 'เวลาโพสต์', 'หมวดหมู่', 'ยอดวิว(ตัวเลข)']
    ws.append(headers)

    # กำหนด font ที่ต้องการ
    font_no_head_thai = Font(name='AnsanamNew')  # ฟอนต์ไม่มีหัวเฉพาะไทย
    red_font = Font(color="FF0000", bold=True)

    # ตั้งค่า column width กว้างพอสมควร
    column_widths = [6, 50, 30, 15, 15, 15, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # ใส่ข้อมูล
    for i, v in enumerate(videos, start=1):
        row = [
            i,
            v.get('title', ''),
            v.get('channel', ''),
            v.get('views', ''),
            v.get('time_posted', ''),
            v.get('category', ''),
            int(v.get('views_num', 0)),
        ]
        ws.append(row)

        # กำหนด style ให้ row ล่าสุด
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            # กึ่งกลางทุกคอลัมน์
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # ฟอนต์ไม่มีหัวสำหรับภาษาไทย (และปกติภาษาอื่น)
            cell.font = font_no_head_thai

        # สีแดง 10 อันดับแรก (ลำดับ <= 10)
        if i <= 10:
            for col in range(1, len(headers)+1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = red_font

    wb.save(filename)
